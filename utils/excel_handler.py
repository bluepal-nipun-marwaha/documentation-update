"""
Excel file handler for section-level editing with formatting preservation.
Supports .xlsx and .xls files using openpyxl.
"""

import io
from copy import copy
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
import structlog

from .document_handlers import BaseDocumentHandler, UpdateSection

logger = structlog.get_logger(__name__)

class ExcelSectionHandler(BaseDocumentHandler):
    """Handler for Excel files with section-level editing."""
    
    def __init__(self):
        super().__init__()
        self.logger = logger
    
    def get_file_extensions(self) -> List[str]:
        """Get supported Excel file extensions."""
        return ['xlsx', 'xls']
    
    def extract_content_with_structure(self, file_bytes: bytes) -> Dict[str, Any]:
        """
        Extract Excel content with structural information.
        
        Args:
            file_bytes: Raw Excel file content as bytes
            
        Returns:
            Dictionary containing sheet structure, data ranges, and formatting metadata
        """
        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
            
            structure = {
                'file_type': 'excel',
                'sheets': {},
                'total_sheets': len(wb.sheetnames),
                'sheet_names': wb.sheetnames
            }
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_structure = self._extract_sheet_structure(sheet)
                structure['sheets'][sheet_name] = sheet_structure
            
            self.logger.info(f"[SUCCESS] Extracted Excel structure: {len(wb.sheetnames)} sheets")
            return structure
            
        except Exception as e:
            self.logger.error(f"[ERROR] Failed to extract Excel structure: {str(e)}")
            return {'file_type': 'excel', 'error': str(e)}
    
    def _extract_sheet_structure(self, sheet) -> Dict[str, Any]:
        """Extract structure from a single Excel sheet."""
        try:
            # Get used range
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Extract headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col)
                headers.append({
                    'column': col,
                    'column_letter': get_column_letter(col),
                    'value': str(cell.value) if cell.value is not None else '',
                    'formatted_value': str(cell.value) if cell.value is not None else ''
                })
            
            # Extract data rows
            data_rows = []
            for row in range(2, max_row + 1):  # Skip header row
                row_data = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_info = {
                        'row': row,
                        'column': col,
                        'column_letter': get_column_letter(col),
                        'value': cell.value,
                        'formatted_value': str(cell.value) if cell.value is not None else '',
                        'formula': cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
                        'cell_id': f"{get_column_letter(col)}{row}",
                        'formatting': self._extract_cell_formatting(cell)
                    }
                    row_data.append(cell_info)
                data_rows.append(row_data)
            
            # Extract merged cells
            merged_ranges = []
            for merged_range in sheet.merged_cells.ranges:
                merged_ranges.append({
                    'range': str(merged_range),
                    'start_row': merged_range.min_row,
                    'end_row': merged_range.max_row,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col
                })
            
            return {
                'sheet_name': sheet.title,
                'max_row': max_row,
                'max_column': max_col,
                'headers': headers,
                'data_rows': data_rows,
                'merged_ranges': merged_ranges,
                'total_data_rows': len(data_rows)
            }
            
        except Exception as e:
            self.logger.error(f"[ERROR] Failed to extract sheet structure: {str(e)}")
            return {'error': str(e)}
    
    def _extract_cell_formatting(self, cell) -> Dict[str, Any]:
        """Extract formatting information from a cell."""
        try:
            formatting = {
                'font': {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'color': str(cell.font.color.rgb) if cell.font.color and cell.font.color.rgb else None
                },
                'fill': {
                    'pattern_type': cell.fill.patternType,
                    'fg_color': str(cell.fill.fgColor.rgb) if cell.fill.fgColor and cell.fill.fgColor.rgb else None,
                    'bg_color': str(cell.fill.bgColor.rgb) if cell.fill.bgColor and cell.fill.bgColor.rgb else None
                },
                'border': {
                    'left': str(cell.border.left.style) if cell.border.left.style else None,
                    'right': str(cell.border.right.style) if cell.border.right.style else None,
                    'top': str(cell.border.top.style) if cell.border.top.style else None,
                    'bottom': str(cell.border.bottom.style) if cell.border.bottom.style else None
                },
                'alignment': {
                    'horizontal': str(cell.alignment.horizontal) if cell.alignment.horizontal else None,
                    'vertical': str(cell.alignment.vertical) if cell.alignment.vertical else None,
                    'wrap_text': cell.alignment.wrap_text
                },
                'number_format': cell.number_format
            }
            return formatting
        except Exception as e:
            self.logger.warning(f"[WARNING] Could not extract cell formatting: {str(e)}")
            return {}
    
    def apply_section_updates(self, file_bytes: bytes, updates: List[UpdateSection]) -> bytes:
        """
        Apply targeted updates to Excel file while preserving formatting.
        
        Args:
            file_bytes: Original Excel file content as bytes
            updates: List of UpdateSection objects describing what to update
            
        Returns:
            Updated Excel file content as bytes
        """
        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
            
            for update in updates:
                self._apply_single_update(wb, update)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            result_bytes = output.getvalue()
            self.logger.info(f"[SUCCESS] Applied {len(updates)} Excel updates: {len(result_bytes)} bytes")
            return result_bytes
            
        except Exception as e:
            self.logger.error(f"[ERROR] Failed to apply Excel updates: {str(e)}")
            return file_bytes  # Return original if update fails
    
    def _apply_single_update(self, workbook, update: UpdateSection):
        """Apply a single update to the workbook."""
        try:
            if update.section_type == 'cell':
                self._update_cell(workbook, update)
            elif update.section_type == 'row':
                self._update_row(workbook, update)
            elif update.section_type == 'sheet':
                self._update_sheet(workbook, update)
            else:
                self.logger.warning(f"[WARNING] Unknown update type: {update.section_type}")
                
        except Exception as e:
            self.logger.error(f"[ERROR] Failed to apply update {update.section_id}: {str(e)}")
    
    def _update_cell(self, workbook, update: UpdateSection):
        """Update a specific cell while preserving formatting."""
        try:
            target = update.target_location
            sheet_name = target.get('sheet_name')
            row = target.get('row')
            col = target.get('column')
            
            if not all([sheet_name, row, col]):
                self.logger.error(f"[ERROR] Invalid cell target: {target}")
                return
            
            sheet = workbook[sheet_name]
            cell = sheet.cell(row=row, column=col)
            
            # Store original formatting
            original_formatting = self._extract_cell_formatting(cell)
            
            # Update cell value
            if update.action == 'replace':
                cell.value = update.new_content
            elif update.action == 'update_formula':
                cell.value = update.new_content  # Formula as string
            
            # Restore formatting
            self._restore_cell_formatting(cell, original_formatting)
            
            self.logger.info(f"[SUCCESS] Updated cell {sheet_name}!{get_column_letter(col)}{row}")
            
        except Exception as e:
            self.logger.error(f"[ERROR] Failed to update cell: {str(e)}")
    
    def _update_row(self, workbook, update: UpdateSection):
        """Update a row of cells while preserving formatting."""
        try:
            target = update.target_location
            sheet_name = target.get('sheet_name')
            row = target.get('row')
            
            if not all([sheet_name, row]):
                self.logger.error(f"[ERROR] Invalid row target: {target}")
                return
            
            sheet = workbook[sheet_name]
            
            if update.action == 'replace_row' and update.new_values:
                # Update each cell in the row
                for col_idx, new_value in enumerate(update.new_values, 1):
                    cell = sheet.cell(row=row, column=col_idx)
                    
                    # Store original formatting
                    original_formatting = self._extract_cell_formatting(cell)
                    
                    # Update value
                    cell.value = new_value
                    
                    # Restore formatting
                    self._restore_cell_formatting(cell, original_formatting)
            
            self.logger.info(f"[SUCCESS] Updated row {row} in sheet {sheet_name}")
            
        except Exception as e:
            self.logger.error(f"[ERROR] Failed to update row: {str(e)}")
    
    def _update_sheet(self, workbook, update: UpdateSection):
        """Update sheet-level properties."""
        try:
            target = update.target_location
            sheet_name = target.get('sheet_name')
            
            if not sheet_name:
                self.logger.error(f"[ERROR] Invalid sheet target: {target}")
                return
            
            sheet = workbook[sheet_name]
            
            if update.action == 'rename_sheet':
                sheet.title = update.new_content
            
            self.logger.info(f"[SUCCESS] Updated sheet {sheet_name}")
            
        except Exception as e:
            self.logger.error(f"[ERROR] Failed to update sheet: {str(e)}")
    
    def _restore_cell_formatting(self, cell, formatting: Dict[str, Any]):
        """Restore formatting to a cell."""
        try:
            # Restore font
            if 'font' in formatting:
                font_data = formatting['font']
                cell.font = Font(
                    name=font_data.get('name'),
                    size=font_data.get('size'),
                    bold=font_data.get('bold'),
                    italic=font_data.get('italic'),
                    color=font_data.get('color')
                )
            
            # Restore fill
            if 'fill' in formatting:
                fill_data = formatting['fill']
                cell.fill = PatternFill(
                    patternType=fill_data.get('pattern_type'),
                    fgColor=fill_data.get('fg_color'),
                    bgColor=fill_data.get('bg_color')
                )
            
            # Restore alignment
            if 'alignment' in formatting:
                align_data = formatting['alignment']
                cell.alignment = Alignment(
                    horizontal=align_data.get('horizontal'),
                    vertical=align_data.get('vertical'),
                    wrap_text=align_data.get('wrap_text')
                )
            
            # Restore number format
            if 'number_format' in formatting:
                cell.number_format = formatting['number_format']
                
        except Exception as e:
            self.logger.warning(f"[WARNING] Could not restore cell formatting: {str(e)}")
    
    def create_cell_update(self, sheet_name: str, row: int, column: int, 
                          new_value: str, reason: str = None) -> UpdateSection:
        """Create a cell update section."""
        return UpdateSection(
            section_id=f"cell_{sheet_name}_{row}_{column}",
            section_type="cell",
            action="replace",
            new_content=new_value,
            target_location={
                'sheet_name': sheet_name,
                'row': row,
                'column': column
            },
            reason=reason
        )
    
    def create_row_update(self, sheet_name: str, row: int, 
                         new_values: List[str], reason: str = None) -> UpdateSection:
        """Create a row update section."""
        return UpdateSection(
            section_id=f"row_{sheet_name}_{row}",
            section_type="row",
            action="replace_row",
            new_values=new_values,
            target_location={
                'sheet_name': sheet_name,
                'row': row
            },
            reason=reason
        )

